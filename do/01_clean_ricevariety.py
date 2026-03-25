#!/usr/bin/env python3
"""
清洗 ricevariety 字段：
1) 统一全角/半角与分隔符；
2) 拆分多品种填报；
3) 纠正常见错别字、缩写、遗漏；
4) 映射到标准品种名与稻作类别，便于后续回归/分组分析。

用法：
python do/01_clean_ricevariety.py \
  --input data_raw/ricevariety_raw.xlsx \
  --output data_clean/survey_ricevariety_clean.xlsx \
  --id-col id
"""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Dict, List, Tuple


MISSING_TOKENS = {"", "na", "n/a", "nan", "0", "不详", "其他", "杂交", "常规", "粳稻", "糯稻"}

# 常见标准化替换（错字/别名/简写 -> 标准表达）
NORMALIZE_RULES: List[Tuple[re.Pattern, str]] = [
    (re.compile(r"甬优\s*4949|勇优\s*4949|甬优4949"), "甬优4949"),
    (re.compile(r"新荣优\s*123|欣荣优$"), "欣荣优123"),
    (re.compile(r"潭两优\s*83|谭两优\s*83|两优83"), "潭两优83"),
    (re.compile(r"荃优\s*822|全优丝苗|全优\s*822"), "荃优822"),
    (re.compile(r"宜香\s*2115|宜香优\s*2115"), "宜香优2115"),
    (re.compile(r"中科发五|中科发\s*5|中禾发\s*5"), "中科发5"),
    (re.compile(r"南粳\s*9108|9108"), "南粳9108"),
    (re.compile(r"南粳\s*5718"), "南粳5718"),
    (re.compile(r"金粳\s*818|粳稻818"), "金粳818"),
    (re.compile(r"龙粳\s*31|(?<!\d)31(?!\d)"), "龙粳31"),
    (re.compile(r"湘早籼\s*45号?|湘早\s*45|湘45"), "湘早籼45"),
    (re.compile(r"湘早籼\s*32号?|湘旱籼\s*32号?"), "湘早籼32"),
    (re.compile(r"中早\s*39|冬早39"), "中早39"),
    (re.compile(r"中早\s*51|中旱51"), "中早51"),
    (re.compile(r"中早\s*83|中旱83"), "中早83"),
    (re.compile(r"早籼\s*617|早稻617|旱米617"), "早籼617"),
    (re.compile(r"陵两优\s*268|株两优819陵两优268"), "陵两优268"),
    (re.compile(r"株两优\s*819|株两优879|株两优189"), "株两优819"),
    (re.compile(r"启两优\s*2216"), "启两优2216"),
    (re.compile(r"野香优航\s*1573.*"), "野香优航1573"),
    (re.compile(r"桃秀优美占"), "桃秀优美珍"),
]

# 品种到类别（可按研究需要继续扩展）
CATEGORY_RULES: List[Tuple[re.Pattern, str]] = [
    (re.compile(r"湘早|中早|早籼|赣早|浙富|中嘉早|中安|株两优|陵两优|兵两优|潭两优|锦两优|启两优"), "早籼稻"),
    (re.compile(r"南粳|金粳|龙粳|绥粳|齐粳|淮稻|粳稻|糯"), "粳稻"),
    (re.compile(r"晚籼|再生稻"), "晚籼稻"),
    (re.compile(r"两优|优|荃优|宜香优|泰优|天优|盛泰优|华盛优|野香优|桃优|隆两优|昱香|华润|西子"), "杂交籼稻"),
    (re.compile(r"香占|丝苗|19香|香稻|美香|留香|桂香|明月"), "优质香稻"),
]


def normalize_text(s: str) -> str:
    s = unicodedata.normalize("NFKC", str(s)).strip().lower()
    s = s.replace("；", ";").replace("，", ",").replace("、", ",").replace("/", ",")
    s = s.replace("（", "(").replace("）", ")")
    s = re.sub(r"[\n\r\t]+", ",", s)
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[\"'“”]+", "", s)
    return s


def split_tokens(s: str) -> List[str]:
    s = re.sub(r"[;|]+", ",", s)
    return [p for p in re.split(r",+", s) if p]


def canonicalize_token(token: str) -> str:
    t = re.sub(r"\(.*?\)", "", token).strip()
    for pat, repl in NORMALIZE_RULES:
        if pat.search(t):
            return repl
    return t.strip(" ,;")


def classify_variety(std: str) -> str:
    if std.lower() in MISSING_TOKENS or std == "":
        return "缺失/未识别"
    for pat, cat in CATEGORY_RULES:
        if pat.search(std):
            return cat
    return "其他/待核验"


def clean_ricevariety(raw: str) -> Tuple[str, str, int]:
    txt = normalize_text(raw)
    tokens = split_tokens(txt)
    std_tokens: List[str] = []
    for tok in tokens:
        c = canonicalize_token(tok)
        if c and c.lower() not in MISSING_TOKENS:
            std_tokens.append(c)

    dedup = list(dict.fromkeys(std_tokens))
    if not dedup:
        return "", "缺失/未识别", 0

    main = dedup[0]
    return main, classify_variety(main), len(dedup)


def _load_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [dict(r) for r in reader]


def _save_csv(path: Path, rows: List[Dict[str, str]], fieldnames: List[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _load_xlsx(path: Path) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 xlsx 需要 openpyxl。请安装后重试，或改用 csv 输入。") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    values = list(ws.values)
    if not values:
        return []
    headers = [str(h) if h is not None else "" for h in values[0]]
    rows = []
    for row in values[1:]:
        rows.append({headers[i]: "" if i >= len(row) or row[i] is None else str(row[i]) for i in range(len(headers))})
    return rows


def _save_xlsx(path: Path, rows: List[Dict[str, str]], fieldnames: List[str]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("写出 xlsx 需要 openpyxl。请安装后重试，或改用 csv 输出。") from exc

    wb = Workbook()
    ws = wb.active
    ws.append(fieldnames)
    for r in rows:
        ws.append([r.get(k, "") for k in fieldnames])
    wb.save(path)


def read_rows(path: Path) -> List[Dict[str, str]]:
    ext = path.suffix.lower()
    if ext == ".csv":
        return _load_csv(path)
    if ext in {".xlsx", ".xls"}:
        return _load_xlsx(path)
    if ext == ".dta":
        raise RuntimeError("当前脚本在无 pandas 环境下不支持 .dta，请先转换为 csv/xlsx。")
    raise ValueError("仅支持 csv/xlsx/dta")


def write_rows(path: Path, rows: List[Dict[str, str]], fieldnames: List[str]) -> None:
    ext = path.suffix.lower()
    if ext == ".csv":
        _save_csv(path, rows, fieldnames)
        return
    if ext in {".xlsx", ".xls"}:
        _save_xlsx(path, rows, fieldnames)
        return
    raise ValueError("--output 仅支持 .csv 或 .xlsx/.xls")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="data_raw/ricevariety_raw.xlsx", help="输入文件（csv/xlsx/dta），默认 data_raw/ricevariety_raw.xlsx")
    parser.add_argument("--output", default="data_clean/survey_ricevariety_clean.xlsx", help="输出文件（csv/xlsx），默认 data_clean/survey_ricevariety_clean.xlsx")
    parser.add_argument("--id-col", default="", help="可选：主键列名")
    args = parser.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    rows = read_rows(in_path)
    if not rows:
        raise RuntimeError("输入数据为空。")
    if "ricevariety" not in rows[0]:
        raise KeyError("未找到列：ricevariety")

    out_rows: List[Dict[str, str]] = []
    counter: Counter[Tuple[str, str]] = Counter()
    for r in rows:
        rr = dict(r)
        main_var, rice_type, n_rep = clean_ricevariety(r.get("ricevariety", ""))
        rr["ricevariety_std"] = main_var
        rr["rice_type_group"] = rice_type
        rr["ricevariety_n_reported"] = str(n_rep)
        rr["ricevariety_multi_flag"] = "1" if n_rep > 1 else "0"
        out_rows.append(rr)
        counter[(rice_type, main_var)] += 1

    fieldnames = list(out_rows[0].keys())
    write_rows(out_path, out_rows, fieldnames)

    summary_path = out_path.with_name(out_path.stem + ("_summary.csv" if out_path.suffix.lower() == ".csv" else "_summary.xlsx"))
    summary_rows = [
        {"rice_type_group": k[0], "ricevariety_std": k[1], "n": str(v)}
        for k, v in sorted(counter.items(), key=lambda kv: (kv[0][0], -kv[1], kv[0][1]))
    ]
    write_rows(summary_path, summary_rows, ["rice_type_group", "ricevariety_std", "n"])

    print(f"[OK] 写出：{out_path}")
    print(f"[OK] 写出：{summary_path}")


if __name__ == "__main__":
    main()
